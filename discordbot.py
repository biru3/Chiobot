from cmath import log
from distutils.sysconfig import PREFIX
import discord
from dotenv import load_dotenv
import os
from multiprocessing.connection import Client
from discord import message
from discord.ext import commands
import asyncio
import random
import openpyxl
from PingPongTool import PingPong

load_dotenv()

intents = discord.Intents.all()
client = discord.Client(intents=intents)
bot = commands.Bot(command_prefix="치오야 ", intents=discord.Intents.all())


@bot.event
async def on_ready():
  print('다음으로 로그인합니다: ')
  print(bot.user.name)
  print('connection was succesful')
  await bot.change_presence(activity=discord.Activity(
    type=discord.ActivityType.listening, name="치오야 도움말"),
                            status=discord.Status.idle)


@bot.command(aliases=['dj','dj설정','접두사설정','음악채널설정','리셋','자동재생','오토','목록초기화','재생목록초기화','가사','자막','지금노래','지금음악','지듣노','지듣음','일시정지','재생','재','이전','되돌리기','목록','재생목록','리스트','플레이리시트','제거','다시재생','재개','검색','찾기','되감기','셔플','랜덤','스킵','넘기기','정지','스탑','볼륨','소리','음악도움말','도움말음악','음악명령어','초대','초대링크','핑','지연시간','상태','봇상태'.'업타임',''])
async def 공지(ctx):
  await ctx.send('')


TOKEN = os.environ['token']
Authorization = os.environ['Auth']
URL = os.environ['URL']

Ping = PingPong(URL, Authorization)
Ping2 = PingPong(URL, Authorization)


@bot.command(aliases=['개발자', '제작자', '만든놈'])
async def 크레딧(ctx):
  embed = discord.Embed(title="크레딧", description="이 봇을 만든놈", color=0xd4b886)
  embed.add_field(name="노예주인", value="아몬드#1734", inline=False)
  embed.add_field(name="코딩노예", value="SPAEER#7411", inline=False)
  embed.set_footer(text="chio")
  await ctx.reply(embed=embed)


@bot.command(aliases=['도움말', '헬프', '도움'])
async def help1(ctx):
  embed = discord.Embed(title="", description="", color=0xd4b886)
  embed.add_field(name="`치오야 배워 [알려줄 말] [뜻]`",
                  value="`저에게 단어를 알려줄 수 있어요!`",
                  inline=False)
  embed.add_field(name="`치오야 음악도움말`", value="`음악 기능의 도움말!`", inline=False)
  embed.add_field(name="`치오야 관리도움말`",
                  value="`서버 관리 명령어를 알려드려요!`",
                  inline=False)
  embed.add_field(name="`치오야 개발자`",
                  value="`아무도 안 궁금해 하는 치오의 개발자를 알려드려요!`",
                  inline=False)
  embed.set_footer(text="Chio")
  await ctx.reply('<@{}>'.format(ctx.message.author.id), embed=embed)


@bot.command(aliases=["관리도움말", "도움말관리"])
async def 도움말4(ctx):
  embed = discord.Embed(title="", description="", color=0xd4b886)
  embed.add_field(name="`치오야 킥 [유저#태그]`", value="`해당 유저를 추방해요!`", inline=False)
  embed.add_field(name="`치오야 밴 [유저#태그]`",
                  value="`해당 유저를 서버에서 차단해요!`",
                  inline=False)
  embed.add_field(name="`치오야 언밴 [유저#태그]`",
                  value="`해당 유저의 차단을 풀어요!`",
                  inline=False)
  embed.add_field(name="`치오야 지워 [삭제할값]`",
                  value="`서버의 채팅을 삭제해요!`",
                  inline=False)
  embed.set_footer(text="Chio")
  await ctx.reply('<@{}>'.format(ctx.message.author.id), embed=embed)


@bot.command(aliases=["삭제", "청소"])
async def 지워(ctx, *, amount):
  i = (ctx.message.author.guild_permissions.administrator)

  if i is True:
    if int(amount) > 100:
      await ctx.reply('100개 이상의 메시지는 지울 수 없어요!')
    else:
      await ctx.channel.purge(limit=int(amount) + 1)
      msg = await ctx.channel.reply('<@{}>, '.format(ctx.message.author.id) +
                                   amount + '개의 메시지를 삭제했습니다!')
      await asyncio.sleep(4)
      await msg.delete()

  if i is False:
    await ctx.channel.reply('<@{}>, '.format(ctx.message.author.id) +
                           '권한이 부족해요..!')


@bot.command(aliases=["추방"])
async def 킥(ctx, *, user_name: discord.Member, reason=None):
  i = (ctx.message.author.guild_permissions.administrator)

  if i is True:
    await user_name.kick(reason=reason)
    await ctx.reply(str(user_name) + "을(를) 추방했어요!")

  if i is False:
    await ctx.channel.reply('<@{}>, '.format(ctx.message.author.id) +
                           '권한이 부족해요..!')


@킥.error
async def 킥_error(ctx, error):
  if isinstance(error, commands.MissingRequiredArgument):
    await ctx.reply("추방할 유저를 넣어주세요!")


@bot.command(aliases=["차단"])
async def 밴(ctx, *, user_name: discord.Member):
  i = (ctx.message.author.guild_permissions.administrator)

  if i is True:
    await user_name.ban()
    await ctx.reply(str(user_name) + "을(를) 차단했어요!")

  if i is False:
    await ctx.channel.reply('<@{}>, '.format(ctx.message.author.id) +
                           '권한이 부족해요..!')


@밴.error
async def 밴_error(ctx, error):
  if isinstance(error, commands.MissingRequiredArgument):
    await ctx.reply("차단할 유저를 넣어주세요!")


@bot.command(aliases=["밴풀기", "차단풀기", "밴풀어", "차단풀어"])
async def 언밴(ctx, *, user_name):
  i = (ctx.message.author.guild_permissions.administrator)

  if i is True:
    banned_users = await ctx.guild.bans()
    member_name, member_discriminator = user_name.split('#')
    for ban_entry in banned_users:
      user = ban_entry.user
      if (user.name, user.discriminator) == (member_name,
                                             member_discriminator):
        await ctx.guild.unban(user)
        await ctx.reply(f"{user.mention}의 차단을 풀었어요!")

  if i is False:
    await ctx.channel.reply('<@{}>, '.format(ctx.message.author.id) +
                           '권한이 부족해요..!')


@지워.error
async def 지워_error(ctx, error):
  if isinstance(error, commands.MissingRequiredArgument):
    await ctx.reply("삭제할 값을 넣어주세요!")


@bot.command()
async def 배워(ctx, learn, *, word):
  wb = openpyxl.load_workbook("memory.xlsx")
  ws = wb.active

  for i in range(1, 10000):
    if ws["A" + str(i)].value == "." or ws["A" + str(i)].value == learn:
      ws["A" + str(i)].value = learn
      ws["B" + str(i)].value = word
      await ctx.reply("추가했어요! 부적절한 단어가 있다면 삭제되거나 수정될 수도 있어요!")
      wb.save("memory.xlsx")
      break


@배워.error
async def 배워_error(ctx, error):
  if isinstance(error, commands.MissingRequiredArgument):
    await ctx.reply('뭘 배울까요?')


@bot.listen()
async def on_command_error(ctx, error):
  if type(error) is commands.errors.CommandNotFound:
    data = await Ping.Pong(ctx.author.id,
                           ctx.message.content[4:],
                           NoTopic=False)
    brand = data['text'][:24]
    오류 = "오류 "
    if data['text'] == "몰?루" or data['text'] == 오류:
      res = ctx.message.content[4:]
      wb = openpyxl.load_workbook("memory.xlsx")
      ws = wb.active

      for i in range(1, 10000):
        val = ws["A" + str(i)].value
        if val == res:
          await ctx.reply(ws["B" + str(i)].value)
          break
      if val == res:
        pass
      else:
        if data['text'] == "몰?루" or data['text'] == 오류:
          response = random.choice(
            ["?", "..네에?", "뭐라구요?", "네?", "몰루?", "으응..", "으음...?"])
          await ctx.reply(response)
    elif brand != "아무말에도 곧잘 대답하는 이 봇은 핑퐁 빌더":
      await ctx.reply(data['text'])
      if data['image']:
        await ctx.reply(data['image'])
    else:
      data2 = await Ping2.Pong(ctx.author.id,
                               ctx.message.content[4:],
                               NoTopic=False)
      await ctx.reply(data2['text'])
      if data2['image']:
        await ctx.reply(data2['image'])


token = os.environ['token']
bot.run(token)
